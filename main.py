import openpyxl as op

from utils.add_or_update_item import add_or_update_item
from utils.file_open import xlsx_filename

file_olta_in = xlsx_filename("olta_in")
file_result_xls = xlsx_filename("result")
# file_sale_by_period = xlsx_filename("sale_by_period")

olta_in_wb = op.load_workbook(file_olta_in, data_only=True)
olta_in_sheet = olta_in_wb.active

result_xls_wb = op.load_workbook(file_result_xls, data_only=True)
result_xls_wb_sheet = result_xls_wb.active

# sale_by_period = op.load_workbook(file_sale_by_period, data_only=True)
# sale_by_period_sheet = sale_by_period.active

item_tyre_dict = {
    "наименование": '',
    "количество_пришло": 0,
    "продано_до_периода": 0,
    "продано_за_период": 0,
    "остаток_на_сейчас": 0
}

list_tyres_dict = []

for i in range(3, olta_in_sheet.max_row):
    tyre_count = olta_in_sheet.cell(row=i, column=3).value
    if tyre_count is not None:
        tyre_name = str(olta_in_sheet.cell(row=i, column=2).value).lower().replace('автошина', '').strip()
        item_tyre_dict['наименование'] = tyre_name
        item_tyre_dict['количество_пришло'] = int(tyre_count)
        list_tyres_dict = add_or_update_item(list_tyres_dict, item_tyre_dict)


print(list_tyres_dict)

result_xls_wb_sheet.cell(row=1, column=1).value = 'Наименование'
result_xls_wb_sheet.cell(row=1, column=2).value = 'Кол-во получено'
i = 2
for item in list_tyres_dict:
    result_xls_wb_sheet.cell(row=i, column=1).value = item['наименование']
    result_xls_wb_sheet.cell(row=i, column=2).value = item['количество_пришло']
    i += 1

# result_xls_wb_sheet.cell(row=i, column=2).value = sum(list_tyres_dict.values())
result_xls_wb.save(file_result_xls)



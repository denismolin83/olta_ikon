import openpyxl as op

from utils.add_or_update_item import add_or_update_item
from utils.file_open import xlsx_filename

file_olta_in = xlsx_filename("olta_in_summer")
file_result_xls = xlsx_filename("result_summer")
file_sale_by_period = xlsx_filename("sale_by_period_summer")

olta_in_wb = op.load_workbook(file_olta_in, data_only=True)
olta_in_sheet = olta_in_wb.active

result_xls_wb = op.load_workbook(file_result_xls, data_only=True)
result_xls_wb_sheet = result_xls_wb.active

sale_by_period = op.load_workbook(file_sale_by_period, data_only=True)
sale_by_period_sheet = sale_by_period.active

item_tyre_dict = {
    "наименование": '',
    "количество_пришло": 0,
    "остаток_на_начало_периода": 0,
    "продано_за_период": 0,
    "остаток_на_сейчас": 0
}

list_tyres_dict = []

# Проходим по экселке с приходом олты и собираем в словарь начальные остатки и остатки на начало периода,
# с учетом продаж
for i in range(3, olta_in_sheet.max_row):
    tyre_count_all = olta_in_sheet.cell(row=i, column=3).value
    tyre_count_start_perion = olta_in_sheet.cell(row=i, column=6).value
    if tyre_count_all is not None:
        tyre_name_sale_by_period = str(olta_in_sheet.cell(row=i, column=2).value).lower().replace('автошина', '').strip()
        item_tyre_dict['наименование'] = tyre_name_sale_by_period
        item_tyre_dict['количество_пришло'] = int(tyre_count_all)
        item_tyre_dict['остаток_на_начало_периода'] = int(tyre_count_start_perion)
        list_tyres_dict = add_or_update_item(list_tyres_dict, item_tyre_dict)


print(list_tyres_dict)

# Проходим по файлу с продажами за период и собираем сколько продано и сколько осталось
for i in range(10, sale_by_period_sheet.max_row):
    tyre_count_sale_by_period = sale_by_period_sheet.cell(row=i, column=14).value
    if tyre_count_sale_by_period is not None:
        tyre_name_sale_by_period = str(sale_by_period_sheet.cell(row=i, column=3).value).lower().replace('автошина', '').strip()
        for item in list_tyres_dict:
            if item['наименование'] == tyre_name_sale_by_period:
                item['продано_за_период'] += int(tyre_count_sale_by_period)
    tyre_count_by_end_period = sale_by_period_sheet.cell(row=i, column=15).value
    if tyre_count_by_end_period is not None:
        tyre_name_by_end_period = str(sale_by_period_sheet.cell(row=i, column=3).value).lower().replace('автошина', '').strip()
        for item in list_tyres_dict:
            if item['наименование'] == tyre_name_by_end_period:
                item['остаток_на_сейчас'] += int(tyre_count_by_end_period)



result_xls_wb_sheet.cell(row=1, column=1).value = 'Наименование'
result_xls_wb_sheet.cell(row=1, column=2).value = 'Кол-во получено всего'
result_xls_wb_sheet.cell(row=1, column=3).value = 'Остаток на начало периода'
result_xls_wb_sheet.cell(row=1, column=4).value = 'Продано за период'
result_xls_wb_sheet.cell(row=1, column=5).value = 'Остаток на конец периода'
i = 2
for item in list_tyres_dict:
    result_xls_wb_sheet.cell(row=i, column=1).value = item['наименование']
    result_xls_wb_sheet.cell(row=i, column=2).value = item['количество_пришло']
    result_xls_wb_sheet.cell(row=i, column=3).value = item['остаток_на_начало_периода']
    result_xls_wb_sheet.cell(row=i, column=4).value = item['продано_за_период']
    result_xls_wb_sheet.cell(row=i, column=5).value = item['остаток_на_сейчас']
    i += 1

# result_xls_wb_sheet.cell(row=i, column=2).value = sum(list_tyres_dict.values())
result_xls_wb_sheet.column_dimensions['A'].width = 43
result_xls_wb.save(file_result_xls)

